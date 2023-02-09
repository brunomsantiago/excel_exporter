import re

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from configuration.sheet import Sheet as SheetConfiguration


def make_table_name(sheet_name: str) -> str:
    table_name = sheet_name.replace(' ', '_').lower()
    table_name = re.sub(r"[^a-zA-Z0-9_]+", '', table_name)
    return table_name


def apply_sheet_configs(ws: Worksheet,
                        ws_config: SheetConfiguration):
    ref = f'B3:{get_column_letter(ws.max_column)}{str(ws.max_row)}'
    table_name = make_table_name(ws_config.sheet_name)
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium1",
                           showFirstColumn=False,
                           showLastColumn=False,
                           showRowStripes=False,
                           showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
