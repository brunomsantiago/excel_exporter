from typing import Dict

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from configuration.column import Column as ColumnConfiguration


def write_columns_header(ws: Worksheet,
                         columns_config: Dict[str, ColumnConfiguration],
                         start_col: int = 2,
                         row_number: int = 3):
    alignment = Alignment(horizontal='left',
                          vertical='center',
                          wrap_text=True)
    font = Font(bold=True, color='FFFFFF')
    for col_number, col_config in enumerate(columns_config.values(),
                                            start=start_col):
        color = col_config.group.background_color.lstrip('#')
        fill = PatternFill(start_color=color,
                           end_color=color,
                           fill_type='solid')
        cell = ws.cell(row=row_number, column=col_number)
        cell.value = col_config.column_name
        cell.fill = fill
        cell.alignment = alignment
        cell.font = font
